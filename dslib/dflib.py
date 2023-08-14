
def get_table_from_excel(workbook_file:str,worksheet_name:str,start_row:int,start_col:int,rows:int,cols:int):
    import openpyxl 
    import pandas as pd
    row_range = range(start_row,start_row+rows,1)
    col_range = range(start_col,start_col+cols,1)
    wb = openpyxl.load_workbook(workbook_file)
    ws = wb[worksheet_name]
    columns=[]
    data = []

    for col in col_range:
        columns.append(ws.cell(row_range[0],col).value)

    row_range = range(start_row+1,start_row+rows,1)

    for row in row_range:
        data_row = []
        for col in col_range:
            data_row.append(ws.cell(row,col).value)
        data.append(data_row)

    return pd.DataFrame(data=data,columns=columns)